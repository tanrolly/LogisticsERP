"""
数据导出 API
支持 Excel 和 CSV 格式导出
"""
from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import csv
import codecs
from flask_login import login_required
from app.models import PurchaseOrder, Order, Inventory, Goods, Warehouse, Supplier
from app.utils.time_helper import beijing_now

bp = Blueprint('export', __name__)


@bp.route('/export/purchase-orders', methods=['GET'])
@login_required
def export_purchase_orders():
    """导出采购订单（Excel/CSV）"""
    format_type = request.args.get('format', 'excel')
    orders = PurchaseOrder.query.order_by(PurchaseOrder.created_at.desc()).all()

    if format_type == 'csv':
        return _export_csv_purchase(orders)
    else:
        return _export_excel_purchase(orders)


@bp.route('/export/transport-orders', methods=['GET'])
@login_required
def export_transport_orders():
    """导出运输订单（Excel/CSV）"""
    format_type = request.args.get('format', 'excel')
    orders = Order.query.order_by(Order.created_at.desc()).all()

    if format_type == 'csv':
        return _export_csv_transport(orders)
    else:
        return _export_excel_transport(orders)


@bp.route('/export/inventory', methods=['GET'])
@login_required
def export_inventory():
    """导出库存数据（Excel/CSV）"""
    format_type = request.args.get('format', 'excel')
    items = Inventory.query.all()

    if format_type == 'csv':
        return _export_csv_inventory(items)
    else:
        return _export_excel_inventory(items)


# ============ Excel 导出 ============

def _export_excel_purchase(orders):
    """采购订单 Excel 导出"""
    wb = Workbook()
    ws = wb.active
    ws.title = "采购订单"

    headers = ['订单编号', '供应商', '订单日期', '交货日期', '总金额', '状态']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for order in orders:
        supplier = Supplier.query.get(order.supplier_id)
        ws.append([
            order.po_no,
            supplier.name if supplier else '',
            order.expected_date.strftime('%Y-%m-%d') if order.expected_date else '',
            '',
            float(order.total_amount) if order.total_amount else 0,
            order.status
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'采购订单_{beijing_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _export_excel_transport(orders):
    """运输订单 Excel 导出"""
    wb = Workbook()
    ws = wb.active
    ws.title = "运输订单"

    headers = ['订单编号', '客户', '起运地', '目的地', '货物名称', '数量', '状态', '预计送达']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for order in orders:
        ws.append([
            order.order_no,
            order.customer.name if order.customer else '',
            order.origin or '',
            order.destination or '',
            order.goods_name or '',
            order.quantity or 0,
            order.status,
            order.plan_arrival.strftime('%Y-%m-%d') if order.plan_arrival else ''
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'运输订单_{beijing_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _export_excel_inventory(items):
    """库存数据 Excel 导出"""
    wb = Workbook()
    ws = wb.active
    ws.title = "库存数据"

    headers = ['商品编号', '商品名称', '仓库', '货位', '库存数量', '更新时间']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for item in items:
        goods = Goods.query.get(item.goods_id)
        warehouse = Warehouse.query.get(item.warehouse_id)
        ws.append([
            goods.code if goods else '',
            goods.name if goods else '',
            warehouse.name if warehouse else '',
            item.location_id or '',
            item.quantity,
            item.updated_at.strftime('%Y-%m-%d %H:%M') if item.updated_at else ''
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'库存数据_{beijing_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============ CSV 导出 ============

def _export_csv_purchase(orders):
    """采购订单 CSV 导出"""
    output = BytesIO()
    output.write(codecs.BOM_UTF8)

    writer = csv.writer(output)
    writer.writerow(['订单编号', '供应商', '订单日期', '交货日期', '总金额', '状态'])

    for order in orders:
        supplier = Supplier.query.get(order.supplier_id)
        writer.writerow([
            order.po_no,
            supplier.name if supplier else '',
            order.expected_date.strftime('%Y-%m-%d') if order.expected_date else '',
            '',
            float(order.total_amount) if order.total_amount else 0,
            order.status
        ])

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'采购订单_{beijing_now().strftime("%Y%m%d_%H%M%S")}.csv',
        mimetype='text/csv'
    )


def _export_csv_transport(orders):
    """运输订单 CSV 导出"""
    output = BytesIO()
    output.write(codecs.BOM_UTF8)

    writer = csv.writer(output)
    writer.writerow(['订单编号', '客户', '起运地', '目的地', '货物名称', '数量', '状态', '预计送达'])

    for order in orders:
        writer.writerow([
            order.order_no,
            order.customer.name if order.customer else '',
            order.origin or '',
            order.destination or '',
            order.goods_name or '',
            order.quantity or 0,
            order.status,
            order.plan_arrival.strftime('%Y-%m-%d') if order.plan_arrival else ''
        ])

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'运输订单_{beijing_now().strftime("%Y%m%d_%H%M%S")}.csv',
        mimetype='text/csv'
    )


def _export_csv_inventory(items):
    """库存数据 CSV 导出"""
    output = BytesIO()
    output.write(codecs.BOM_UTF8)

    writer = csv.writer(output)
    writer.writerow(['商品编号', '商品名称', '仓库', '货位', '库存数量', '更新时间'])

    for item in items:
        goods = Goods.query.get(item.goods_id)
        warehouse = Warehouse.query.get(item.warehouse_id)
        writer.writerow([
            goods.code if goods else '',
            goods.name if goods else '',
            warehouse.name if warehouse else '',
            item.location_id or '',
            item.quantity,
            item.updated_at.strftime('%Y-%m-%d %H:%M') if item.updated_at else ''
        ])

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'库存数据_{beijing_now().strftime("%Y%m%d_%H%M%S")}.csv',
        mimetype='text/csv'
    )
